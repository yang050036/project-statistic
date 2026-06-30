import json
import requests
import sys

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

API_BASE_URL = "http://localhost:3001/api"


def parse_args():
    args = sys.argv[1:]
    parsed = {}
    i = 0
    while i < len(args):
        if args[i].startswith('--'):
            key = args[i].replace('--', '')
            if i + 1 < len(args) and not args[i + 1].startswith('--'):
                value = args[i + 1]
                i += 2
            else:
                value = 'true'
                i += 1
            parsed[key] = value
        else:
            i += 1
    return parsed


def init_api_url(args):
    global API_BASE_URL
    if 'api' in args:
        API_BASE_URL = args['api']
    elif 'host' in args or 'port' in args:
        host = args.get('host', 'localhost')
        port = args.get('port', '3001')
        API_BASE_URL = f"http://{host}:{port}/api"
    print(f"API地址: {API_BASE_URL}")


def push_task(task_data):
    url = f"{API_BASE_URL}/tasks"
    try:
        response = requests.post(url, json=task_data)
        response.raise_for_status()
        print(f"[OK] 成功推送任务: {task_data['customNumber']}")
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"[ERR] 推送失败: {e}")
        return None


def push_tasks(tasks):
    results = []
    for task in tasks:
        result = push_task(task)
        if result:
            results.append(result)
    print(f"\n批量推送完成，成功 {len(results)} 条")
    return results


def get_all_tasks():
    url = f"{API_BASE_URL}/tasks"
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"获取数据失败: {e}")
        return []


def update_task(task_id, updates):
    url = f"{API_BASE_URL}/tasks/{task_id}"
    try:
        response = requests.put(url, json=updates)
        response.raise_for_status()
        print(f"[OK] 成功更新任务: {task_id}")
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"[ERR] 更新失败: {e}")
        return None


def delete_task(task_id):
    url = f"{API_BASE_URL}/tasks/{task_id}"
    try:
        response = requests.delete(url)
        response.raise_for_status()
        print(f"[OK] 成功删除任务: {task_id}")
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"[ERR] 删除失败: {e}")
        return None


def push_from_excel(file_path):
    if not EXCEL_SUPPORT:
        print("[ERR] 请先安装 openpyxl: pip install openpyxl")
        return []

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        tasks = []
        headers = []
        
        for row in ws.iter_rows(max_row=1):
            headers = [cell.value for cell in row]
        
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = header.strip().lower()
                if '项目号' in header_lower or 'project' in header_lower:
                    header_map['projectNumber'] = i
                elif '人员' in header_lower or 'person' in header_lower:
                    header_map['person'] = i
                elif '个性化号' in header_lower or 'customnumber' in header_lower:
                    header_map['customNumber'] = i
                elif '个性化内容' in header_lower or 'customcontent' in header_lower:
                    header_map['customContent'] = i
                elif '交付路径' in header_lower or 'deliverypath' in header_lower:
                    header_map['deliveryPath'] = i
                elif '消耗人时' in header_lower or 'hours' in header_lower:
                    header_map['hours'] = i
                elif '交付时间' in header_lower or 'deliverytime' in header_lower:
                    header_map['deliveryTime'] = i
        
        required_fields = ['projectNumber', 'person', 'customNumber', 'customContent', 'deliveryPath', 'hours', 'deliveryTime']
        missing_fields = [f for f in required_fields if f not in header_map]
        if missing_fields:
            print(f"[ERR] Excel文件缺少必要列: {', '.join(missing_fields)}")
            return []
        
        for row in ws.iter_rows(min_row=2):
            task = {}
            for field, idx in header_map.items():
                value = row[idx].value
                if field == 'hours':
                    value = float(value) if value else 0
                elif field == 'deliveryTime':
                    value = str(value) if value else ''
                task[field] = value
            if task['projectNumber']:
                tasks.append(task)
        
        print(f"从Excel文件读取到 {len(tasks)} 条数据")
        return push_tasks(tasks)
    except Exception as e:
        print(f"[ERR] 读取Excel文件失败: {e}")
        return []


def push_from_args(args):
    task_data = {
        'projectNumber': args.get('projectNumber', ''),
        'person': args.get('person', ''),
        'customNumber': args.get('customNumber', ''),
        'customContent': args.get('customContent', ''),
        'deliveryPath': args.get('deliveryPath', ''),
        'hours': float(args.get('hours', '0')),
        'deliveryTime': args.get('deliveryTime', '')
    }
    
    if not task_data['projectNumber'] or not task_data['person']:
        print("[ERR] 缺少必要参数: --projectNumber 和 --person 为必填")
        return None
    
    return push_task(task_data)


def interactive_mode():
    sample_task = {
        "projectNumber": "P006",
        "person": "钱七",
        "customNumber": "C016",
        "customContent": "系统性能优化",
        "deliveryPath": "/docs/P006/优化报告.docx",
        "hours": 20,
        "deliveryTime": "2024-06-01"
    }

    print("=== 项目统计数据推送脚本 ===")
    print(f"当前API地址: {API_BASE_URL}")
    print("")
    print("使用方法:")
    print("  python push_data.py --push --projectNumber P001 --person 张三 ...")
    print("  python push_data.py --excel <Excel文件路径>")
    print("")
    print("指定服务器(可选):")
    print("  python push_data.py --api http://192.168.1.100:3001/api")
    print("  python push_data.py --host 192.168.1.100 --port 3001")
    print("")
    print("默认API地址: http://localhost:3001/api")
    print("")
    print("1. 推送单个任务")
    print("2. 批量推送任务")
    print("3. 从Excel文件导入")
    print("4. 获取所有任务")
    print("5. 更新任务")
    print("6. 删除任务")
    print("7. 退出")

    while True:
        choice = input("\n请输入选择: ")

        if choice == "1":
            push_task(sample_task)

        elif choice == "2":
            batch_tasks = [
                {
                    "projectNumber": "P006",
                    "person": "钱七",
                    "customNumber": "C017",
                    "customContent": "安全漏洞修复",
                    "deliveryPath": "/docs/P006/安全报告.docx",
                    "hours": 8,
                    "deliveryTime": "2024-06-05"
                },
                {
                    "projectNumber": "P006",
                    "person": "张三",
                    "customNumber": "C018",
                    "customContent": "文档更新",
                    "deliveryPath": "/docs/P006/更新文档.docx",
                    "hours": 4,
                    "deliveryTime": "2024-06-10"
                }
            ]
            push_tasks(batch_tasks)

        elif choice == "3":
            file_path = input("请输入Excel文件路径: ")
            push_from_excel(file_path)

        elif choice == "4":
            tasks = get_all_tasks()
            print(f"\n共 {len(tasks)} 条任务:")
            for task in tasks:
                print(f"  - {task['projectNumber']} | {task['person']} | {task['customNumber']} | {task['hours']}h")

        elif choice == "5":
            task_id = input("请输入任务ID: ")
            updates = {"hours": 25}
            update_task(task_id, updates)

        elif choice == "6":
            task_id = input("请输入任务ID: ")
            delete_task(task_id)

        elif choice == "7":
            print("退出脚本")
            break

        else:
            print("无效选择，请重新输入")


if __name__ == "__main__":
    args = parse_args()
    init_api_url(args)

    if args.get('push'):
        push_from_args(args)
    elif args.get('excel'):
        push_from_excel(args['excel'])
    else:
        interactive_mode()
